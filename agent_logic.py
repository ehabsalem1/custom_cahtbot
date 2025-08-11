import os
import json
import dotenv
from openai import OpenAI
from pathlib import Path
from openpyxl import Workbook, load_workbook

from langchain_community.vectorstores import FAISS
from langchain_community.embeddings import OpenAIEmbeddings
from langchain_community.chat_models import ChatOpenAI
from langchain.chains.question_answering import load_qa_chain
from langchain_community.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter

dotenv.load_dotenv()
client = OpenAI()

EXCEL_PATH = Path("data/tickets.xlsx")

# --- Ticket logging tool ---

def write_to_excel(name, phone, issue):
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Issue"])
    ws.append([name, phone, issue])
    wb.save(EXCEL_PATH)

ticket_tool = {
    "type": "function",
    "function": {
        "name": "create_ticket",
        "description": "Log an issue into the Excel ticket file",
        "parameters": {
            "type": "object",
            "properties": {
                "name": {"type": "string"},
                "phone": {"type": "string"},
                "issue": {"type": "string"}
            },
            "required": ["name", "phone", "issue"]
        }
    }
}

SYSTEM_PROMPT = """
You are an IT helpdesk assistant. Help users solve problems using
knowledge base, and if they give you an issue with name and phone, log
it using create_ticket.
"""

# --- PDF Vectorstore setup ---

VECTORSTORE_DIR = "vectorstore"
PDF_PATH = "data\knowledge\content.pdf"  #PDF path here

def build_vectorstore_from_pdf(pdf_path):
    print("Loading PDF and building vectorstore...")
    loader = PyPDFLoader(pdf_path)
    docs = loader.load()

    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
    docs = text_splitter.split_documents(docs)

    embeddings = OpenAIEmbeddings()
    db = FAISS.from_documents(docs, embeddings)
    db.save_local(VECTORSTORE_DIR)
    print("Vectorstore built and saved.")

# Build vectorstore if not exists
if not os.path.exists(VECTORSTORE_DIR):
    build_vectorstore_from_pdf(PDF_PATH)

embeddings = OpenAIEmbeddings()
db = FAISS.load_local(VECTORSTORE_DIR, embeddings, allow_dangerous_deserialization=True)
retriever = db.as_retriever()

llm = ChatOpenAI(model_name="gpt-4o")
chain = load_qa_chain(llm, chain_type="stuff")

def search_knowledge_base(query):
    docs = retriever.get_relevant_documents(query)
    if not docs:
        return "Sorry, no answer found in the knowledge base."
    return chain.run(input_documents=docs, question=query)

def ask_agent(user_input, history=[]):
    # Convert internal history format to OpenAI chat messages
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    for turn in history:
        messages.append({"role": "user", "content": turn["user"]})
        messages.append({"role": "assistant", "content": turn["bot"]})
    messages.append({"role": "user", "content": user_input})

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=messages,
        tools=[ticket_tool],
        tool_choice="auto"
    )
    msg = response.choices[0].message
    if msg.tool_calls:
        args = json.loads(msg.tool_calls[0].function.arguments)
        write_to_excel(args["name"], args["phone"], args["issue"])
        return f"✅ Ticket saved for {args['name']}! We’ll contact them at {args['phone']}."
    
    # If no tool call, try knowledge base
    kb_answer = search_knowledge_base(user_input)
    if kb_answer:
        return kb_answer

    return msg.content
